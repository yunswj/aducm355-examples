#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据导出工具模块
"""

import csv
import json
import xml.etree.ElementTree as ET
from datetime import datetime
import numpy as np
from PyQt5.QtWidgets import QMessageBox, QFileDialog
import os

class DataExporter:
    """数据导出器"""
    
    def __init__(self):
        self.supported_formats = {
            'csv': 'CSV文件 (*.csv)',
            'json': 'JSON文件 (*.json)',
            'xml': 'XML文件 (*.xml)',
            'txt': '文本文件 (*.txt)',
            'xlsx': 'Excel文件 (*.xlsx)'
        }
    
    def export_data(self, data_points, parent_widget=None, default_filename=None):
        """
        导出数据
        
        Args:
            data_points: 数据点列表 [(voltage, current, timestamp), ...]
            parent_widget: 父窗口组件
            default_filename: 默认文件名
        
        Returns:
            bool: 导出是否成功
        """
        if not data_points:
            if parent_widget:
                QMessageBox.information(parent_widget, "提示", "没有数据可导出")
            return False
        
        # 构建文件过滤器
        filters = ";;".join(self.supported_formats.values()) + ";;所有文件 (*.*)"
        
        # 默认文件名
        if not default_filename:
            default_filename = f"CV_Data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        # 选择保存路径
        file_path, selected_filter = QFileDialog.getSaveFileName(
            parent_widget, "导出数据", default_filename, filters
        )
        
        if not file_path:
            return False
        
        try:
            # 根据文件扩展名选择导出格式
            ext = os.path.splitext(file_path)[1].lower()
            
            if ext == '.csv':
                self.export_to_csv(data_points, file_path)
            elif ext == '.json':
                self.export_to_json(data_points, file_path)
            elif ext == '.xml':
                self.export_to_xml(data_points, file_path)
            elif ext == '.txt':
                self.export_to_txt(data_points, file_path)
            elif ext == '.xlsx':
                self.export_to_excel(data_points, file_path)
            else:
                # 默认导出为CSV
                self.export_to_csv(data_points, file_path)
            
            if parent_widget:
                QMessageBox.information(parent_widget, "成功", f"数据已导出到:\n{file_path}")
            
            return True
            
        except Exception as e:
            if parent_widget:
                QMessageBox.critical(parent_widget, "错误", f"导出数据失败:\n{str(e)}")
            return False
    
    def export_to_csv(self, data_points, file_path):
        """导出到CSV文件"""
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # 写入标题行
            writer.writerow(['序号', '电压(V)', '电流(A)', '时间戳', '时间'])
            
            # 写入数据
            for i, (voltage, current, timestamp) in enumerate(data_points):
                dt = datetime.fromtimestamp(timestamp / 1000.0)
                time_str = dt.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                writer.writerow([i + 1, voltage, current, timestamp, time_str])
    
    def export_to_json(self, data_points, file_path):
        """导出到JSON文件"""
        data = {
            'metadata': {
                'export_time': datetime.now().isoformat(),
                'data_count': len(data_points),
                'measurement_type': 'Cyclic Voltammetry',
                'format_version': '1.0'
            },
            'statistics': self.calculate_statistics(data_points),
            'data': []
        }
        
        for i, (voltage, current, timestamp) in enumerate(data_points):
            dt = datetime.fromtimestamp(timestamp / 1000.0)
            data['data'].append({
                'index': i + 1,
                'voltage': voltage,
                'current': current,
                'timestamp': timestamp,
                'datetime': dt.isoformat()
            })
        
        with open(file_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(data, jsonfile, indent=2, ensure_ascii=False)
    
    def export_to_xml(self, data_points, file_path):
        """导出到XML文件"""
        root = ET.Element("CyclicVoltammetryData")
        
        # 元数据
        metadata = ET.SubElement(root, "Metadata")
        ET.SubElement(metadata, "ExportTime").text = datetime.now().isoformat()
        ET.SubElement(metadata, "DataCount").text = str(len(data_points))
        ET.SubElement(metadata, "MeasurementType").text = "Cyclic Voltammetry"
        ET.SubElement(metadata, "FormatVersion").text = "1.0"
        
        # 统计信息
        stats = self.calculate_statistics(data_points)
        statistics = ET.SubElement(root, "Statistics")
        for key, value in stats.items():
            ET.SubElement(statistics, key.replace('_', '')).text = str(value)
        
        # 数据点
        data_element = ET.SubElement(root, "DataPoints")
        
        for i, (voltage, current, timestamp) in enumerate(data_points):
            point = ET.SubElement(data_element, "DataPoint")
            point.set("index", str(i + 1))
            
            ET.SubElement(point, "Voltage").text = str(voltage)
            ET.SubElement(point, "Current").text = str(current)
            ET.SubElement(point, "Timestamp").text = str(timestamp)
            
            dt = datetime.fromtimestamp(timestamp / 1000.0)
            ET.SubElement(point, "DateTime").text = dt.isoformat()
        
        # 写入文件
        tree = ET.ElementTree(root)
        tree.write(file_path, encoding='utf-8', xml_declaration=True)
    
    def export_to_txt(self, data_points, file_path):
        """导出到文本文件"""
        with open(file_path, 'w', encoding='utf-8') as txtfile:
            # 写入标题
            txtfile.write("循环伏安法测量数据\n")
            txtfile.write("=" * 50 + "\n")
            txtfile.write(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            txtfile.write(f"数据点数: {len(data_points)}\n")
            txtfile.write("=" * 50 + "\n\n")
            
            # 统计信息
            stats = self.calculate_statistics(data_points)
            txtfile.write("统计信息:\n")
            txtfile.write("-" * 30 + "\n")
            txtfile.write(f"电压范围: {stats['voltage_min']:.6f} ~ {stats['voltage_max']:.6f} V\n")
            txtfile.write(f"电流范围: {stats['current_min']:.9f} ~ {stats['current_max']:.9f} A\n")
            txtfile.write(f"平均电流: {stats['current_mean']:.9f} A\n")
            txtfile.write(f"电流标准差: {stats['current_std']:.9f} A\n")
            txtfile.write(f"测量时长: {stats['duration']:.3f} s\n")
            txtfile.write("\n")
            
            # 数据表格
            txtfile.write("数据详情:\n")
            txtfile.write("-" * 80 + "\n")
            txtfile.write(f"{'序号':<6} {'电压(V)':<12} {'电流(A)':<15} {'时间戳':<15} {'时间':<20}\n")
            txtfile.write("-" * 80 + "\n")
            
            for i, (voltage, current, timestamp) in enumerate(data_points):
                dt = datetime.fromtimestamp(timestamp / 1000.0)
                time_str = dt.strftime("%H:%M:%S.%f")[:-3]
                txtfile.write(f"{i+1:<6} {voltage:<12.6f} {current:<15.9f} {timestamp:<15.0f} {time_str:<20}\n")
    
    def export_to_excel(self, data_points, file_path):
        """导出到Excel文件"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("需要安装openpyxl库才能导出Excel文件: pip install openpyxl")
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        
        # 数据工作表
        ws_data = wb.active
        ws_data.title = "测量数据"
        
        # 设置标题行
        headers = ['序号', '电压(V)', '电流(A)', '时间戳', '时间']
        for col, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 写入数据
        for i, (voltage, current, timestamp) in enumerate(data_points, 2):
            ws_data.cell(row=i, column=1, value=i-1)
            ws_data.cell(row=i, column=2, value=voltage)
            ws_data.cell(row=i, column=3, value=current)
            ws_data.cell(row=i, column=4, value=timestamp)
            
            dt = datetime.fromtimestamp(timestamp / 1000.0)
            ws_data.cell(row=i, column=5, value=dt.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3])
        
        # 自动调整列宽
        for col in range(1, 6):
            column_letter = get_column_letter(col)
            ws_data.column_dimensions[column_letter].auto_size = True
        
        # 统计信息工作表
        ws_stats = wb.create_sheet("统计信息")
        stats = self.calculate_statistics(data_points)
        
        # 写入统计信息
        stats_data = [
            ["项目", "值", "单位"],
            ["数据点数", len(data_points), "个"],
            ["电压最小值", stats['voltage_min'], "V"],
            ["电压最大值", stats['voltage_max'], "V"],
            ["电压范围", stats['voltage_range'], "V"],
            ["电流最小值", stats['current_min'], "A"],
            ["电流最大值", stats['current_max'], "A"],
            ["电流范围", stats['current_range'], "A"],
            ["电流平均值", stats['current_mean'], "A"],
            ["电流标准差", stats['current_std'], "A"],
            ["测量时长", stats['duration'], "s"],
            ["导出时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ""]
        ]
        
        for row, data in enumerate(stats_data, 1):
            for col, value in enumerate(data, 1):
                cell = ws_stats.cell(row=row, column=col, value=value)
                if row == 1:  # 标题行
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 自动调整列宽
        for col in range(1, 4):
            column_letter = get_column_letter(col)
            ws_stats.column_dimensions[column_letter].auto_size = True
        
        # 保存文件
        wb.save(file_path)
    
    def calculate_statistics(self, data_points):
        """计算统计信息"""
        if not data_points:
            return {}
        
        voltages = [point[0] for point in data_points]
        currents = [point[1] for point in data_points]
        timestamps = [point[2] for point in data_points]
        
        stats = {
            'voltage_min': min(voltages),
            'voltage_max': max(voltages),
            'voltage_range': max(voltages) - min(voltages),
            'voltage_mean': np.mean(voltages),
            'voltage_std': np.std(voltages),
            'current_min': min(currents),
            'current_max': max(currents),
            'current_range': max(currents) - min(currents),
            'current_mean': np.mean(currents),
            'current_std': np.std(currents),
            'duration': (max(timestamps) - min(timestamps)) / 1000.0 if len(timestamps) > 1 else 0
        }
        
        return stats
    
    def get_supported_formats(self):
        """获取支持的导出格式"""
        return list(self.supported_formats.keys())
    
    def get_format_filter(self, format_key):
        """获取指定格式的文件过滤器"""
        return self.supported_formats.get(format_key, "所有文件 (*.*)")

# 全局导出器实例
data_exporter = DataExporter()